import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')
wb = openpyxl.load_workbook(r'C:\Users\jay\Downloads\hana_ai_excel_practice_instructor_v4.xlsx', data_only=True)
ws1 = wb['결과작성_STEP1_보고용집계']
ws2 = wb['결과작성_STEP2_오류찾기']
ws_r = wb['거래원장_RAW']

with open('verify_v4.txt','w',encoding='utf-8') as f:
    f.write("=== STEP1 배너 ===\n")
    f.write(str(ws1['A1'].value)+'\n\n')
    f.write("=== STEP1 직원별 (insert 후 row6부터) ===\n")
    for i in range(20):
        r = 6+i
        vals = [ws1.cell(r,c).value for c in range(1,8)]
        f.write("|".join(str(v) for v in vals)+'\n')
    f.write("\n=== STEP2 배너 ===\n")
    f.write(str(ws2['A1'].value)+'\n\n')
    f.write("=== STEP2 발견건수 (row6,7) ===\n")
    f.write(f"C6={ws2.cell(6,3).value} C7={ws2.cell(7,3).value}\n\n")
    f.write("=== K열 오류 행 ===\n")
    for row in ws_r.iter_rows(min_row=5, values_only=True):
        if row[10]: f.write(f"{row[0]} {row[1]} {row[2]}: K='{row[10]}'\n")

print(open('verify_v4.txt',encoding='utf-8').read())
