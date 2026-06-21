"""v4 파일 기준 STEP1/STEP2 정답 계산"""
import sys
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

PATH = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v4.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)
ws_raw = wb['거래원장_RAW']
ws_m   = wb['직원마스터']

VALID_PROD  = {'주식','펀드','채권','ELS','RP'}
VALID_TRADE = {'매수','매도','가입','해지'}

# 직원마스터
master = {}
for row in ws_m.iter_rows(min_row=4, values_only=True):
    if row[0] and str(row[0]).startswith('E'):
        master[row[0]] = (row[1], row[2])

rows = list(ws_raw.iter_rows(min_row=5, values_only=True))

# STEP1 집계: 처리상태=정상 기준
emp_normal_cnt = defaultdict(int)
emp_normal_amt = defaultdict(int)
emp_normal_fee = defaultdict(int)
emp_err_cnt    = defaultdict(int)

for r in rows:
    date, branch, emp, cust, prod, trade, amt, rate, fee, status = r[:10]
    if not emp: continue
    amt_i = int(amt) if amt else 0
    fee_i = int(fee) if fee else 0
    if status == '정상':
        emp_normal_cnt[emp] += 1
        emp_normal_amt[emp] += amt_i
        emp_normal_fee[emp] += fee_i
    else:
        emp_err_cnt[emp] += 1

print("=== STEP1 직원별 집계 정답 ===")
EMPLOYEES = [f'E{str(i).zfill(3)}' for i in range(1,21)]
with open('v4_answers.txt', 'w', encoding='utf-8') as f:
    f.write("=== STEP1 직원별 집계 정답 ===\n")
    for e in EMPLOYEES:
        name, branch = master.get(e, ('?','?'))
        line = f"{e}|{name}|{branch}|{emp_normal_cnt[e]}|{emp_normal_amt[e]}|{emp_normal_fee[e]}|{emp_err_cnt[e]}"
        print(line)
        f.write(line + '\n')

    print("\n=== STEP2 오류 행 ===")
    f.write("\n=== STEP2 오류 행 ===\n")
    for ri, r in enumerate(rows):
        if r[9] != '정상': continue
        date, branch, emp, cust, prod, trade, amt, rate, fee, status = r[:10]
        try: amt_f = float(amt) if amt is not None else None
        except: amt_f = None
        errs = []
        if amt_f is None or amt_f <= 0: errs.append('거래금액 오류')
        if prod not in VALID_PROD: errs.append('상품유형 오류')
        if errs:
            row_no = ri + 5
            line = f"row={row_no}|{date}|{branch}|{emp}|{cust}|prod={prod}|amt={amt}|→{errs[0]}"
            print(line)
            f.write(line + '\n')

print("\n✅ v4_answers.txt 저장 완료")
