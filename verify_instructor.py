"""강사용 정답 파일 검증"""
import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

PATH = r'C:\Users\jay\Downloads\hana_ai_excel_practice_instructor_v3.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)

print("=== 시트 목록 ===")
for s in wb.sheetnames:
    tc = wb[s].sheet_properties.tabColor
    print(f"  {s}  tabColor={tc}")

print("\n=== STEP1 지점별 (행 2가 헤더 삽입으로 밀림 → 데이터는 6행부터) ===")
ws1 = wb['결과작성_STEP1_보고용집계']
# 강사용 주석 행(1행) 삽입으로 모든 행이 +1 됨
print("  A1:", ws1['A1'].value)
for r in range(6, 11):
    branch = ws1.cell(r, 1).value
    cnt    = ws1.cell(r, 2).value
    amt    = ws1.cell(r, 3).value
    fee    = ws1.cell(r, 4).value
    err    = ws1.cell(r, 5).value
    print(f"  {branch}: 건수={cnt}, 금액={amt:,}, 수수료={fee:,}, 취소오류={err}")

print("\n=== STEP1 상품유형별 ===")
prods = ['주식','펀드','채권','ELS','RP']
for i, p in enumerate(prods):
    r = 6 + i
    cnt  = ws1.cell(r, 9).value
    amt  = ws1.cell(r, 10).value
    fee  = ws1.cell(r, 11).value
    rat  = ws1.cell(r, 12).value
    print(f"  {p}: 건수={cnt}, 금액={amt:,}, 수수료={fee:,}, 비중={rat}")

print("\n=== STEP1 직원별 (E001~E005 확인) ===")
for i in range(5):
    r = 14 + i   # row1 삽입 후: E001=14, E002=15, ...
    eid  = ws1.cell(r, 1).value
    name = ws1.cell(r, 2).value
    bran = ws1.cell(r, 3).value
    cnt  = ws1.cell(r, 4).value
    fee  = ws1.cell(r, 5).value
    err  = ws1.cell(r, 6).value
    fee_fmt = f"{fee:,}" if isinstance(fee, (int,float)) else str(fee)
    print(f"  {eid} {name} {bran}: 정상건수={cnt}, 수수료={fee_fmt}, 취소오류={err}")

print("\n=== STEP2 발견 건수 ===")
ws2 = wb['결과작성_STEP2_오류찾기']
print("  A1:", ws2['A1'].value)
types = ['거래금액 오류','수수료율 오류','수수료금액 오류','상품유형 오류','거래유형 오류']
for i, t in enumerate(types):
    cnt = ws2.cell(6+i, 3).value  # row1 삽입 후 원래 C5→C6
    print(f"  {t}: {cnt}건")

print("\n=== 거래원장_RAW K열 오류 행 ===")
ws_r = wb['거래원장_RAW']
for row in ws_r.iter_rows(min_row=5, values_only=True):
    if row[10]:
        print(f"  {row[0]} {row[1]} {row[2]} {row[3]}: K='{row[10]}'")
