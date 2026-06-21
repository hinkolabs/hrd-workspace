"""
STEP2 리디자인: 원본 RAW 데이터를 추출 후 새 워크북으로 빌드
"""
import sys, copy
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_simple_errorfind.xlsx'
DST = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v2.xlsx'

# ── 원본에서 RAW 데이터 추출 ─────────────────────────────────
src_wb = openpyxl.load_workbook(SRC, data_only=True)
raw_rows = list(src_wb['거래원장_RAW'].iter_rows(min_row=1, values_only=True))
master_rows = list(src_wb['직원마스터'].iter_rows(min_row=1, values_only=True))
src_wb.close()

# ── 색상 팔레트 ──────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

BLUE_DARK   = fill("1F3864")
BLUE_MID    = fill("2F5496")
BLUE_LIGHT  = fill("D9E1F2")
YELLOW_FILL = fill("FFF2CC")
GREEN_FILL  = fill("E2EFDA")
ORANGE_FILL = fill("FCE4D6")
GRAY_FILL   = fill("F2F2F2")

def font(bold=False, color="1F3864", size=10, italic=False):
    return Font(name="맑은 고딕", bold=bold, color=color, size=size, italic=italic)

WHITE_BOLD = font(True, "FFFFFF", 11)
DARK_BOLD  = font(True, "1F3864", 10)
DARK_NORM  = font(False, "1F3864", 10)
RED_BOLD   = font(True, "C00000", 10)
GRAY_NORM  = font(False, "595959", 9, italic=True)
GREEN_BOLD = font(True, "375623", 10)

CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, value, bg=BLUE_DARK, fg=WHITE_BOLD, align=CENTER, h=26):
    c = ws.cell(row, col, value)
    c.fill = bg; c.font = fg; c.alignment = align; c.border = thin_border()
    ws.row_dimensions[row].height = h
    return c

def cell(ws, row, col, value, bg=None, fg=None, align=LEFT, bdr=True):
    c = ws.cell(row, col, value)
    if bg:  c.fill = bg
    if fg:  c.font = fg
    c.alignment = align
    if bdr: c.border = thin_border()
    return c

# ── 새 워크북 생성 ───────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # 기본 Sheet 제거

# ════════════════════════════════════════════════════════════
# 시트 1: 00_수강생_README
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("00_수강생_README")
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 76

hdr(ws, 1, 1, "구분");  hdr(ws, 1, 2, "내용")
data = [
    ("실습 목적",  "AI와 함께 거래 데이터를 정리하고, 시스템이 '정상'이라고 표시한 데이터에서 실제 오류를 찾아냅니다."),
    ("실습 흐름",  "STEP1 보고용 집계표 완성  →  STEP2 이상 거래 탐지"),
    ("STEP1 핵심", "처리상태=정상 기준으로 지점별·상품별·직원별 집계.\nAI에게 수식을 요청하고 결과를 직접 검증하세요."),
    ("STEP2 핵심", "처리상태가 '정상'이어도 수치값이 이상한 행을 찾습니다.\n"
                  "확인 항목: ① 거래금액 이상  ② 수수료율 이상  ③ 수수료금액 계산 오류\n"
                  "AI에게 먼저 '어떤 이상이 있을 수 있는지' 물어보고, 수식을 만들고, 결과를 검증하세요."),
    ("주의사항",   "AI가 만든 수식을 그대로 믿지 마세요. 샘플 행을 눈으로 확인하는 것이 핵심입니다."),
    ("데이터 안내","실제 고객 데이터가 아닌 가상 거래 원장입니다. 의도적으로 심어진 오류가 있습니다."),
]
for i, (a, b) in enumerate(data, 2):
    cell(ws, i, 1, a, fg=DARK_BOLD, align=CENTER)
    cell(ws, i, 2, b, fg=DARK_NORM, align=LEFT_TOP)
    ws.row_dimensions[i].height = 40


# ════════════════════════════════════════════════════════════
# 시트 2: 01_실습흐름
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("01_실습흐름")
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 36
ws.column_dimensions['E'].width = 28

ws.merge_cells('A1:E1')
hdr(ws, 1, 1, "3교시 실습 흐름", h=28)

for j, h in enumerate(["시간", "단계", "내용", "수강생 활동", "강사 포인트"], 1):
    hdr(ws, 2, j, h, h=24)

flow = [
    ("0~5분",    "도입",   "파일 수령 강의 아님 — AI 업무처리 실습 설명", "파일 열기, RAW 구조 확인", "답보다 조건 설명·검증이 핵심"),
    ("5~25분",   "STEP1",  "보고용 집계표 완성",             "AI 프롬프트로 수식 생성 후 입력", "기존 STEP2→STEP1 진행"),
    ("25~30분",  "STEP2\n탐색", "AI에게 오류 패턴 탐색 요청",   "탐색 프롬프트 복붙 후 AI 응답 검토", "조건을 먼저 묻게 유도"),
    ("30~40분",  "STEP2\n수식", "오류유형 K열 수식 생성",        "수식 생성 프롬프트 → K열 입력",      "처리상태=정상 조건 확인"),
    ("40~45분",  "STEP2\n검증", "AI 결과 반박·검증",             "검증 프롬프트로 반례 찾기",           "AI가 틀린 케이스 체험"),
    ("45~50분",  "마무리",  "핵심 메시지 정리",               "체크리스트 확인 후 제출",             "AI는 도구, 판단은 사람"),
]
for i, row in enumerate(flow, 3):
    bg = BLUE_LIGHT if i % 2 == 0 else None
    for j, v in enumerate(row, 1):
        cell(ws, i, j, v, bg=bg, fg=DARK_NORM, align=CENTER if j <= 2 else LEFT)
    ws.row_dimensions[i].height = 36


# ════════════════════════════════════════════════════════════
# 시트 3: 거래원장_RAW  (원본 그대로 복사)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("거래원장_RAW")
col_widths = [12, 10, 8, 8, 10, 10, 16, 10, 14, 10, 22, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r_idx, row in enumerate(raw_rows, 1):
    for c_idx, val in enumerate(row, 1):
        c = ws.cell(r_idx, c_idx, val)
        if r_idx == 1:
            pass  # 원본 주석 행
        elif r_idx == 2:
            pass
        elif r_idx == 3:  # 헤더
            c.fill = BLUE_DARK; c.font = WHITE_BOLD; c.alignment = CENTER; c.border = thin_border()
        else:
            c.font = DARK_NORM
            c.alignment = CENTER
            c.border = thin_border()
            if r_idx % 2 == 0: c.fill = GRAY_FILL
    if r_idx == 3:
        ws.row_dimensions[r_idx].height = 24
    else:
        ws.row_dimensions[r_idx].height = 18

# 헤더 행 강조
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 22
ws['A1'].font = font(True, "C00000", 10)
ws['A2'].font = font(False, "595959", 9, italic=True)


# ════════════════════════════════════════════════════════════
# 시트 4: 직원마스터 (원본 복사)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("직원마스터")
for i, w in enumerate([10, 12, 10, 8, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for r_idx, row in enumerate(master_rows, 1):
    for c_idx, val in enumerate(row, 1):
        c = ws.cell(r_idx, c_idx, val)
        if r_idx == 1:
            c.fill = BLUE_DARK; c.font = WHITE_BOLD; c.alignment = CENTER
        elif r_idx == 2:
            c.font = GRAY_NORM; c.alignment = LEFT
        elif r_idx == 3:
            c.fill = BLUE_MID; c.font = WHITE_BOLD; c.alignment = CENTER; c.border = thin_border()
        else:
            c.font = DARK_NORM; c.alignment = CENTER; c.border = thin_border()
            if r_idx % 2 == 0: c.fill = GRAY_FILL
    ws.row_dimensions[r_idx].height = 20


# ════════════════════════════════════════════════════════════
# 시트 5: AI작업_캔버스
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("AI작업_캔버스")
for i, w in enumerate([10, 24, 40, 22, 32, 28, 16], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells('A1:G1')
hdr(ws, 1, 1, "AI 작업 캔버스  —  각 단계에서 AI에게 맡긴 작업과 검증 내용을 기록하세요", h=28)

for j, h in enumerate(["단계", "AI에게 맡긴 작업", "데이터 설명 / 힌트", "조건", "AI 응답 확인사항", "내가 검증한 내용", "최종 판단"], 1):
    hdr(ws, 2, j, h, h=24)

canvas = [
    ("STEP1", "보고용 집계표 완성", "거래원장_RAW 4행~, 직원마스터 4행~",
     "처리상태 기준: 정상/취소+오류 구분", "", "", ""),
    ("STEP2\n탐색", "오류 패턴 추론 요청", "G 거래금액, H 수수료율, I 수수료금액, J 처리상태",
     "처리상태=정상인 행만", "", "", ""),
    ("STEP2\n수식", "K열 오류유형 수식 생성", "탐색 결과로 확정된 조건 사용",
     "① 거래금액≤0  ② 수수료율>0.5%  ③ 수수료금액 불일치", "", "", ""),
    ("STEP2\n검증", "AI 수식 반례 검증", "거래금액=0이지만 정상인 케이스? 반올림 오차?",
     "취소·오류 행 제외됐는지 확인", "", "", ""),
]
for i, row in enumerate(canvas, 3):
    bg = BLUE_LIGHT if i % 2 == 0 else None
    for j, v in enumerate(row, 1):
        cell(ws, i, j, v, bg=bg, fg=DARK_BOLD if j == 1 else DARK_NORM,
             align=CENTER if j == 1 else LEFT_TOP)
    ws.row_dimensions[i].height = 52


# ════════════════════════════════════════════════════════════
# 시트 6: 복붙용_AI프롬프트  (탐색→생성→검증 3단계)
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("복붙용_AI프롬프트")
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 82
ws.column_dimensions['C'].width = 36
ws.column_dimensions['D'].width = 18

ws.merge_cells('A1:D1')
hdr(ws, 1, 1, "복붙용 AI 프롬프트  —  아래 텍스트를 복사해서 AI 채팅창에 붙여넣으세요", h=28)

for j, h in enumerate(["구분", "프롬프트 (복사해서 붙여넣기)", "AI 응답 확인 포인트", "비고"], 1):
    hdr(ws, 2, j, h, h=24)

prompts = [
    (
        "STEP1\n보고용 집계",
        ("아래 거래원장 시트 구조를 참고해서 보고용 집계표를 만들어줘.\n\n"
         "컬럼: A 거래일자, B 지점코드, C 직원ID, D 고객ID, E 상품유형, "
         "F 거래유형, G 거래금액, H 수수료율, I 수수료금액, J 처리상태\n\n"
         "만들어야 할 표:\n"
         "① 지점별(B001~B005): 처리상태=정상인 거래건수·거래금액합계·수수료합계, 취소+오류 건수\n"
         "② 상품유형별(주식/펀드/채권/ELS/RP): 처리상태=정상인 건수·금액·수수료, 전체 대비 비중\n"
         "③ 직원별(E001~E020): 직원마스터 시트에서 직원명 VLOOKUP, 정상 건수·수수료, 취소+오류 건수\n\n"
         "결과작성_STEP1_보고용집계 시트에 들어갈 Excel 수식으로 만들어줘. "
         "RAW 데이터는 거래원장_RAW 시트 4행부터 시작해."),
        "정상/취소·오류 조건이 분리됐는지 확인",
        "기존 STEP2→STEP1"
    ),
    (
        "STEP2\n① 탐색",
        ("거래 데이터에서 처리상태가 '정상'으로 표시됐지만 실제 수치값에 문제가 있을 수 있는 경우를 추론해줘.\n\n"
         "컬럼 구조: G 거래금액(원, 양수여야 함), H 수수료율(소수점, 예: 0.0014), "
         "I 수수료금액(원), J 처리상태\n\n"
         "처리상태=정상인 행에서 발생할 수 있는 수치 이상을 3가지 이내로 추론해줘. "
         "각 조건이 실무적으로 말이 되는지도 같이 판단해줘."),
        "AI 제안 중 누락되거나\n이상한 조건 체크",
        "조건을 AI에게 먼저 묻기"
    ),
    (
        "STEP2\n② 수식 생성",
        ("거래원장_RAW 시트에서 처리상태(J열)가 '정상'인 행만 대상으로,\n"
         "아래 3가지 오류 조건을 K열에 표시하는 Excel 수식을 만들어줘.\n\n"
         "오류 조건:\n"
         "1. 거래금액(G) ≤ 0  →  '거래금액 오류'\n"
         "2. 수수료율(H) > 0.005 (0.5% 초과) 또는 ≤ 0  →  '수수료율 오류'\n"
         "3. 수수료금액(I) ≠ 거래금액(G) × 수수료율(H)  (1원 초과 차이)  →  '수수료금액 오류'\n\n"
         "여러 조건이 해당하면 첫 번째 조건만 표시해도 돼. 조건 없으면 빈 칸. 4행부터 적용."),
        "처리상태=정상 조건이\n수식에 포함됐는지 확인",
        "탐색 후 조건 확정"
    ),
    (
        "STEP2\n③ 검증",
        ("방금 만든 K열 수식을 검증해줘.\n\n"
         "아래 케이스를 직접 확인해줘:\n"
         "- 거래금액이 0인데 수수료금액이 있는 행  →  오류로 잡히는가?\n"
         "- 수수료금액이 음수인 행  →  오류로 잡히는가?\n"
         "- 처리상태가 '취소'·'오류'인 행  →  K열이 빈 칸인가? (정상만 검사해야 함)\n"
         "- 수수료금액이 계산값과 정확히 1원 차이인 행  →  오류로 잡히지 않는가? (반올림 허용)\n\n"
         "수식에 문제가 있으면 수정안도 같이 줘."),
        "AI가 스스로 반례를 찾는지\n확인하는 것이 핵심",
        "AI 결과 검증"
    ),
]
for i, (구분, 프롬, 체크, 비고) in enumerate(prompts, 3):
    bg = BLUE_LIGHT if i % 2 == 0 else None
    cell(ws, i, 1, 구분, bg=bg, fg=DARK_BOLD, align=CENTER)
    cell(ws, i, 2, 프롬, bg=bg, fg=DARK_NORM, align=LEFT_TOP)
    cell(ws, i, 3, 체크, bg=bg, fg=GRAY_NORM, align=LEFT_TOP)
    cell(ws, i, 4, 비고, bg=bg, fg=GRAY_NORM, align=CENTER)
    ws.row_dimensions[i].height = 100


# ════════════════════════════════════════════════════════════
# 시트 7: 결과작성_STEP1_보고용집계
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("결과작성_STEP1_보고용집계")
col_ws = [12, 16, 18, 18, 16, 14, 4, 12, 16, 18, 18, 14, 14]
for i, w in enumerate(col_ws, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells('A1:M1')
hdr(ws, 1, 1, "STEP1 결과 작성: 보고용 집계표", h=28)

ws.merge_cells('A2:M2')
c = ws.cell(2, 1, "처리상태=정상 기준 집계. 취소·오류 건수는 별도. 직원명은 직원마스터 VLOOKUP.")
c.fill = YELLOW_FILL; c.font = DARK_NORM; c.alignment = LEFT
ws.row_dimensions[2].height = 20

# 지점별 집계
ws.merge_cells('A3:F3')
hdr(ws, 3, 1, "① 지점별 집계", bg=BLUE_MID, h=22)

for j, h in enumerate(["지점코드", "정상 거래건수", "정상 거래금액", "정상 수수료 합계", "취소·오류 건수", "비고"], 1):
    hdr(ws, 4, j, h, h=24)

for i, branch in enumerate(["B001","B002","B003","B004","B005"], 5):
    bg = GRAY_FILL if i % 2 == 0 else None
    cell(ws, i, 1, branch, bg=bg, fg=DARK_BOLD, align=CENTER)
    for j in range(2, 7):
        cell(ws, i, j, "", bg=bg, fg=DARK_NORM, align=CENTER)
    ws.row_dimensions[i].height = 20

# 상품유형별 집계 (H~M)
ws.merge_cells('H3:M3')
hdr(ws, 3, 8, "② 상품유형별 집계", bg=BLUE_MID, h=22)

for j, h in enumerate(["상품유형", "정상 거래건수", "정상 거래금액", "정상 수수료 합계", "거래 비중", "비고"], 1):
    hdr(ws, 4, j+7, h, h=24)

for i, prod in enumerate(["주식","펀드","채권","ELS","RP"], 5):
    bg = GRAY_FILL if i % 2 == 0 else None
    cell(ws, i, 8, prod, bg=bg, fg=DARK_BOLD, align=CENTER)
    for j in range(9, 14):
        cell(ws, i, j, "", bg=bg, fg=DARK_NORM, align=CENTER)

# 직원별 처리 현황
ws.merge_cells('A11:G11')
hdr(ws, 11, 1, "③ 직원별 처리 현황  (직원명: 직원마스터 시트 VLOOKUP)", bg=BLUE_MID, h=22)

for j, h in enumerate(["직원ID", "직원명", "지점코드", "정상 거래건수", "정상 수수료 합계", "취소·오류 건수", "비고"], 1):
    hdr(ws, 12, j, h, h=24)

for i, emp in enumerate([f"E{str(n).zfill(3)}" for n in range(1, 21)], 13):
    bg = GRAY_FILL if i % 2 == 0 else None
    cell(ws, i, 1, emp, bg=bg, fg=DARK_BOLD, align=CENTER)
    for j in range(2, 8):
        cell(ws, i, j, "", bg=bg, fg=DARK_NORM, align=CENTER)
    ws.row_dimensions[i].height = 18


# ════════════════════════════════════════════════════════════
# 시트 8: 결과작성_STEP2_오류찾기
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("결과작성_STEP2_오류찾기")
for i, w in enumerate([14, 10, 10, 10, 12, 10, 16, 10, 14, 22, 16], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells('A1:K1')
hdr(ws2, 1, 1, "STEP2 결과 작성: 이상 거래 탐지  (처리상태=정상인 행 대상)", h=28)

ws2.merge_cells('A2:K2')
c = ws2.cell(2, 1, "AI가 K열에 오류유형을 표시하면, 아래 요약표의 발견 건수를 직접 확인·입력하세요.")
c.fill = YELLOW_FILL; c.font = DARK_NORM; c.alignment = LEFT
ws2.row_dimensions[2].height = 20

# 오류 유형 요약표
ws2.merge_cells('A3:D3')
hdr(ws2, 3, 1, "오류 유형별 요약", bg=BLUE_MID, h=22)

for j, h in enumerate(["오류 유형", "판단 기준", "발견 건수", "비고"], 1):
    hdr(ws2, 4, j, h, h=24)

summary = [
    ("거래금액 이상",   "처리상태=정상  AND  거래금액 ≤ 0",                                 "", ""),
    ("수수료율 이상",   "처리상태=정상  AND  수수료율 > 0.5% 또는 ≤ 0",                      "", "0.5% = 0.005"),
    ("수수료금액 오류", "처리상태=정상  AND  수수료금액 ≠ 거래금액×수수료율  (±1원 허용)",    "", "반올림 허용"),
]
for i, (유형, 기준, 건수, 비고) in enumerate(summary, 5):
    bg = GREEN_FILL if i % 2 == 1 else None
    cell(ws2, i, 1, 유형, bg=bg, fg=DARK_BOLD, align=CENTER)
    cell(ws2, i, 2, 기준, bg=bg, fg=DARK_NORM, align=LEFT)
    cell(ws2, i, 3, 건수, bg=bg, fg=RED_BOLD, align=CENTER)
    cell(ws2, i, 4, 비고, bg=bg, fg=GRAY_NORM, align=CENTER)
    ws2.row_dimensions[i].height = 26

# 오류 행 목록
ws2.merge_cells('A9:K9')
hdr(ws2, 9, 1, "오류 행 목록  —  K열에 오류유형이 표시된 행을 여기에 정리하세요", bg=BLUE_MID, h=24)

list_hdrs = ["거래일자","지점","직원ID","고객ID","상품유형","거래유형","거래금액","수수료율","수수료금액","오류유형","정정처리상태"]
for j, h in enumerate(list_hdrs, 1):
    hdr(ws2, 10, j, h, h=24)

for i in range(11, 22):
    bg = GRAY_FILL if i % 2 == 0 else None
    for j in range(1, 12):
        cell(ws2, i, j, "", bg=bg)
    ws2.row_dimensions[i].height = 18

# AI 검증 메모 영역
ws2.merge_cells('A23:K23')
c = ws2.cell(23, 1, "AI 검증 메모  —  AI가 잘못 분류한 케이스, 또는 AI가 못 잡은 케이스를 기록하세요")
c.fill = ORANGE_FILL; c.font = RED_BOLD; c.alignment = LEFT
ws2.row_dimensions[23].height = 24

for i in range(24, 27):
    ws2.merge_cells(f'A{i}:K{i}')
    ws2[f'A{i}'].border = thin_border()
    ws2.row_dimensions[i].height = 22


# ════════════════════════════════════════════════════════════
# 시트 9: 제출전_체크리스트
# ════════════════════════════════════════════════════════════
ws = wb.create_sheet("제출전_체크리스트")
ws.column_dimensions['A'].width = 54
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 14

ws.merge_cells('A1:D1')
hdr(ws, 1, 1, "제출 전 체크리스트", h=28)

for j, h in enumerate(["항목", "체크", "메모", "강사 확인"], 1):
    hdr(ws, 2, j, h, h=24)

checks = [
    "STEP1: 처리상태=정상 기준으로 집계했는가?",
    "STEP1: 직원명을 직원마스터에서 VLOOKUP으로 가져왔는가?",
    "STEP1: 취소·오류 건수를 별도 집계했는가?",
    "STEP2: 탐색 프롬프트로 AI에게 오류 패턴을 먼저 물어봤는가?",
    "STEP2: 처리상태=정상 조건이 K열 수식에 포함됐는가?",
    "STEP2: 수수료금액 계산 오류를 반올림 허용(±1원)으로 체크했는가?",
    "STEP2: AI 수식 결과를 샘플 행으로 직접 눈으로 검증했는가?",
    "AI 검증 메모에 AI가 틀린 케이스를 한 건 이상 기록했는가?",
]
for i, item in enumerate(checks, 3):
    bg = GRAY_FILL if i % 2 == 0 else None
    cell(ws, i, 1, item, bg=bg, fg=DARK_NORM, align=LEFT)
    cell(ws, i, 2, "", bg=bg, align=CENTER)
    cell(ws, i, 3, "", bg=bg)
    cell(ws, i, 4, "", bg=bg, align=CENTER)
    ws.row_dimensions[i].height = 22


# ── 탭 색상 설정 ─────────────────────────────────────────────
tab_colors = {
    "00_수강생_README":           "1F3864",
    "01_실습흐름":                "2F5496",
    "거래원장_RAW":               "C00000",
    "직원마스터":                  "375623",
    "AI작업_캔버스":              "7030A0",
    "복붙용_AI프롬프트":          "FF9900",
    "결과작성_STEP1_보고용집계":  "0070C0",
    "결과작성_STEP2_오류찾기":    "0070C0",
    "제출전_체크리스트":          "595959",
}
for name, color in tab_colors.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = color

wb.save(DST)
print(f"저장 완료 → {DST}")
print()
print("변경 내역:")
print("  복붙용_AI프롬프트  : STEP2 탐색→수식생성→검증 3단계 프롬프트")
print("  결과작성_STEP2     : 3가지 오류유형 요약 + AI 검증 메모 영역")
print("  AI작업_캔버스      : 탐색 단계 추가")
print("  제출전_체크리스트  : AI 검증 관련 항목 추가")
print("  거래원장_RAW       : 원본 데이터 보존 (오류 5건 그대로)")
