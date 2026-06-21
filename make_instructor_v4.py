"""v4 학생 파일 → 강사용 정답 파일 생성"""
import shutil, sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SRC  = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v4.xlsx'
DEST = r'C:\Users\jay\Downloads\hana_ai_excel_practice_instructor_v4.xlsx'

shutil.copy2(SRC, DEST)
wb = openpyxl.load_workbook(DEST)

# ── 스타일 ─────────────────────────────────────────────────────────────────────
ANSWER_FILL  = PatternFill("solid", fgColor="FFF2CC")
ANS_FONT_NUM = Font(bold=True, color="7F4600", size=9)
ANS_FONT_STR = Font(bold=True, color="1F3864", size=9)
HEADER_FILL  = PatternFill("solid", fgColor="C00000")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=12)
THIN         = Side(style="thin", color="BFBFBF")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
center = Alignment(horizontal='center', vertical='center')
right  = Alignment(horizontal='right',  vertical='center')

def fill_ans(ws, row, col, value, is_num=True, fmt=None):
    c = ws.cell(row, col)
    c.value      = value
    c.fill       = ANSWER_FILL
    c.font       = ANS_FONT_NUM if is_num else ANS_FONT_STR
    c.alignment  = right if is_num else center
    c.border     = THIN_BORDER
    if fmt: c.number_format = fmt

# ── 정답 데이터 ────────────────────────────────────────────────────────────────
EMP_DATA = [
    ("E001","김민준","B001", 8, 1_550_500_000,  2_753_630, 0),
    ("E002","이서연","B002", 6, 1_418_400_000,  3_547_980, 1),
    ("E003","박도윤","B003",11, 2_583_200_000,  7_068_900, 2),
    ("E004","최하은","B004",10, 1_769_600_000,  3_062_510, 0),
    ("E005","정지호","B005", 8, 1_930_600_000,  4_841_060, 2),
    ("E006","강서준","B001", 8, 2_654_100_000,  4_014_180, 3),
    ("E007","조지민","B002",10, 1_280_400_000,  3_015_520, 0),
    ("E008","윤현우","B003",10, 2_835_800_000,  6_170_200, 3),
    ("E009","장수아","B004", 8, 2_208_000_000,  2_890_410, 0),
    ("E010","임유진","B005", 7, 1_301_200_000,  2_054_780, 2),
    ("E011","한시우","B001", 4,   152_100_000,    362_020, 2),
    ("E012","오예준","B002",12, 2_335_700_000,  4_971_880, 2),
    ("E013","서지우","B003", 8, 1_683_600_000,  3_992_350, 2),
    ("E014","신하린","B004",11, 2_344_500_000,  4_192_220, 1),
    ("E015","권민재","B005", 6, 2_083_100_000,  4_285_510, 0),
    ("E016","황나은","B001", 6, 1_861_400_000,  3_546_230, 2),
    ("E017","안준호","B002",13, 3_442_000_000,  6_339_080, 0),
    ("E018","송다은","B003", 8, 1_816_900_000,  4_923_520, 3),
    ("E019","류태윤","B004", 6, 1_973_900_000,  4_310_400, 3),
    ("E020","홍채원","B005", 8, 1_925_600_000,  4_019_350, 3),
]

# ── STEP1: 직원별 집계표 ────────────────────────────────────────────────────────
ws1 = wb['결과작성_STEP1_보고용집계']

# 강사용 배너 (row 1 위에 삽입)
ws1.insert_rows(1)
ws1.merge_cells('A1:G1')
hdr = ws1['A1']
hdr.value     = '★ 강사용 정답 파일 — 배포 금지 ★'
hdr.fill      = HEADER_FILL
hdr.font      = HEADER_FONT
hdr.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 26

# 데이터 행: insert_rows(1) 후 E001 = row6, E002 = row7 ...
for i, (eid, name, branch, cnt, amt, fee, err) in enumerate(EMP_DATA):
    r = 6 + i  # 원래 row5 → 삽입 후 row6
    fill_ans(ws1, r, 2, name,   is_num=False)          # B: 직원명
    fill_ans(ws1, r, 3, branch, is_num=False)          # C: 지점코드
    fill_ans(ws1, r, 4, cnt,    fmt='0')               # D: 정상건수
    fill_ans(ws1, r, 5, amt,    fmt='#,##0')           # E: 정상거래금액
    fill_ans(ws1, r, 6, fee,    fmt='#,##0')           # F: 정상수수료합계
    fill_ans(ws1, r, 7, err,    fmt='0')               # G: 취소오류건수

ws1.sheet_properties.tabColor = "FF6B35"
print("✓ STEP1 완료")

# ── STEP2: 발견 건수 ────────────────────────────────────────────────────────────
ws2 = wb['결과작성_STEP2_오류찾기']

ws2.insert_rows(1)
ws2.merge_cells('A1:D1')
hdr2 = ws2['A1']
hdr2.value     = '★ 강사용 정답 파일 — 배포 금지 ★'
hdr2.fill      = HEADER_FILL
hdr2.font      = HEADER_FONT
hdr2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 26

# 원래 C5, C6 → 삽입 후 C6, C7
fill_ans(ws2, 6, 3, 5, fmt='0')   # 거래금액 오류: 5건
fill_ans(ws2, 7, 3, 3, fmt='0')   # 상품유형 오류: 3건

ws2.sheet_properties.tabColor = "FF6B35"
print("✓ STEP2 완료")

# ── 거래원장_RAW K열: 8개 오류 행 ──────────────────────────────────────────────
ws_raw = wb['거래원장_RAW']

ERR_ROWS = {
    (20250101, 'B002', 'E007', 'C032'): '거래금액 오류',
    (20250102, 'B001', 'E011', 'C041'): '거래금액 오류',
    (20250103, 'B002', 'E012', 'C018'): '상품유형 오류',
    (20250103, 'B004', 'E009', 'C075'): '거래금액 오류',
    (20250103, 'B004', 'E014', 'C073'): '상품유형 오류',
    (20250104, 'B004', 'E004', 'C066'): '거래금액 오류',
    (20250104, 'B005', 'E005', 'C016'): '거래금액 오류',
    (20250122, 'B002', 'E017', 'C065'): '상품유형 오류',
}

K_FILL = PatternFill("solid", fgColor="FFE4B5")
K_FONT = Font(bold=True, color="8B0000", size=9)

found = 0
for row_idx, row_obj in enumerate(ws_raw.iter_rows(min_row=5)):
    vals = [c.value for c in row_obj]
    key = (vals[0], vals[1], vals[2], vals[3])
    if key in ERR_ROWS:
        c = ws_raw.cell(row_obj[0].row, 11)
        c.value = ERR_ROWS[key]
        c.fill  = K_FILL
        c.font  = K_FONT
        print(f"  K{row_obj[0].row} = '{ERR_ROWS[key]}'  ({key[0]} {key[1]})")
        found += 1

ws_raw.sheet_properties.tabColor = "C00000"
print(f"✓ K열 완료 ({found}/8건)")

# ── 저장 ──────────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f"\n✅ 저장: {DEST}")
