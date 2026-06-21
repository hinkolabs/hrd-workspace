import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_simplified_errorcheck.xlsx',
    data_only=True
)
ws = wb['거래원장_RAW']

# 수정/보존 대상 행 정확한 엑셀 행번호 출력
targets = {
    # (date, branch, empid, custid): 설명
    (20250104, 'B004', 'E004', 'C066'): 'KEEP-거래금액0',
    (20250109, 'B002', 'E002', 'C021'): 'KEEP-수수료율1.5%',
    (20250117, 'B005', 'E015', 'C007'): 'FIX-수수료금액불일치',
    (20250122, 'B002', 'E017', 'C065'): 'KEEP-ETF',
    (20250129, 'B001', 'E001', 'C040'): 'FIX-거래유형정정',
    (20250224, 'B002', 'E012', 'C035'): 'FIX-수수료금액음수',
    (20250306, 'B001', 'E001', 'C079'): 'FIX-거래금액음수',
    (20250319, 'B003', 'E008', 'C071'): 'FIX-수수료율None',
}

for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
    key = (row[0].value, row[1].value, row[2].value, row[3].value)
    if key in targets:
        r = row[0].row
        v = [c.value for c in row[:12]]
        print(f"[엑셀행{r}] {targets[key]}")
        print(f"  날짜={v[0]} 지점={v[1]} 직원={v[2]} 고객={v[3]}")
        print(f"  상품={v[4]} 거래={v[5]} 금액={v[6]} 수수료율={v[7]} 수수료금액={v[8]} 상태={v[9]}")
        print()
