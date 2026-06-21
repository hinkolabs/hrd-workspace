import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

XLS = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v4.xlsx'
wb = openpyxl.load_workbook(XLS, data_only=True)
print('=== 시트 목록 ===')
for name in wb.sheetnames:
    print(f'  {name}')

for shname in wb.sheetnames:
    ws = wb[shname]
    print(f'\n\n=== 시트: [{shname}] (행={ws.max_row}, 열={ws.max_column}) ===')
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
        vals = [str(c) if c is not None else '' for c in row]
        line = ' | '.join(vals).rstrip(' |')
        if line.replace('|','').strip():
            print(f'  {line}')
