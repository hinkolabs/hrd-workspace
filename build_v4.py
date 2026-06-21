"""
v3_fixed → v4 변환
변경 사항:
  1. RAW: 수수료율 오류 행 정상 복구
          거래금액 오류 4건 추가 (총 5건)
          상품유형 오류 2건 추가 (총 3건)
  2. STEP1: 3개 집계표 → 1개 통합 직원별 집계표
  3. STEP2: 5개 오류유형 → 2개 (거래금액/상품유형)
"""
import shutil, sys, copy
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import copy

SRC  = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_fixed.xlsx'
DEST = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v4.xlsx'

shutil.copy2(SRC, DEST)
wb = openpyxl.load_workbook(DEST)
ws_raw = wb['거래원장_RAW']
ws_m   = wb['직원마스터']

# ──────────────────────────────────────────────────────────────────────────────
# 0. 기존 RAW 데이터 파악
# ──────────────────────────────────────────────────────────────────────────────
VALID_PROD  = {'주식','펀드','채권','ELS','RP'}

all_rows = []
for row_obj in ws_raw.iter_rows(min_row=5):
    vals = [c.value for c in row_obj]
    if vals[0] is None: continue
    all_rows.append({'row': row_obj[0].row, 'vals': vals})

print(f"총 데이터 행 수: {len(all_rows)}")

# 처리상태=정상 행만
normal_rows = [r for r in all_rows if r['vals'][9] == '정상']
print(f"처리상태=정상 행: {len(normal_rows)}")

# ── 오류 행 선정 기준: 정상 행 중 다양한 지점/날짜에서 선택 ──────────────────

# 현재 오류 행 (touch하지 말 것):
EXISTING_ERR_KEYS = {
    (20250104, 'B004', 'E004', 'C066'),  # 거래금액=0 (이미 오류)
    (20250109, 'B002', 'E002', 'C021'),  # 수수료율=1.5% (제거 대상)
    (20250122, 'B002', 'E017', 'C065'),  # 상품유형=ETF (이미 오류)
}

# 후보 행: 정상이면서 기존 오류 행이 아닌 것
candidates = []
for r in normal_rows:
    v = r['vals']
    key = (v[0], v[1], v[2], v[3])
    if key not in EXISTING_ERR_KEYS:
        candidates.append(r)

# 지점 분포 확인
by_branch = defaultdict(list)
for c in candidates:
    by_branch[c['vals'][1]].append(c)

print("\n지점별 후보 수:")
for b in sorted(by_branch):
    print(f"  {b}: {len(by_branch[b])}개")

# ── 거래금액 오류 추가 4건 선정: 지점 골고루, 다른 날짜 ──────────────────────
# B001, B003, B004, B005에서 하나씩 (B002는 이미 수수료율 오류 → 정상화 후 거래금액으로)
# 간단하게: 각 지점별로 처음 나오는 후보 선택
amt_err_targets = []
used_branches = set()
for r in candidates:
    branch = r['vals'][1]
    if branch not in used_branches and branch not in ('B002',):  # B002는 수수료율 오류 복구로 대신
        amt_err_targets.append(r)
        used_branches.add(branch)
    if len(amt_err_targets) == 3:
        break

# B002에서도 하나 선택 (수수료율 오류 행 제외한 다른 행)
for r in by_branch['B002']:
    v = r['vals']
    key = (v[0], v[1], v[2], v[3])
    if key not in EXISTING_ERR_KEYS:
        amt_err_targets.append(r)
        break

print(f"\n거래금액 오류 추가 대상 ({len(amt_err_targets)}건):")
for r in amt_err_targets:
    v = r['vals']
    print(f"  row={r['row']} {v[0]} {v[1]} {v[2]} {v[3]} 상품={v[4]} 금액={v[6]}")

# ── 상품유형 오류 추가 2건 선정 ──────────────────────────────────────────────
prod_err_targets = []
used_branches2 = set()
used_rows = {r['row'] for r in amt_err_targets}
for r in candidates:
    branch = r['vals'][1]
    if r['row'] not in used_rows and branch not in used_branches2:
        prod_err_targets.append(r)
        used_branches2.add(branch)
    if len(prod_err_targets) == 2:
        break

print(f"\n상품유형 오류 추가 대상 ({len(prod_err_targets)}건):")
for r in prod_err_targets:
    v = r['vals']
    print(f"  row={r['row']} {v[0]} {v[1]} {v[2]} {v[3]} 상품={v[4]}")

# ──────────────────────────────────────────────────────────────────────────────
# 1. RAW 수정
# ──────────────────────────────────────────────────────────────────────────────

def get_raw_row(date, branch, emp, cust):
    for r in all_rows:
        v = r['vals']
        if v[0]==date and v[1]==branch and v[2]==emp and v[3]==cust:
            return r['row']
    return None

# 수수료율 오류 행 복구: 수수료율을 0.003으로, 수수료금액도 재계산
fix_row = get_raw_row(20250109, 'B002', 'E002', 'C021')
if fix_row:
    normal_rate = 0.003
    amt = ws_raw.cell(fix_row, 7).value  # G = 거래금액
    new_fee = round(amt * normal_rate) if amt else 0
    ws_raw.cell(fix_row, 8).value = normal_rate     # H = 수수료율
    ws_raw.cell(fix_row, 9).value = new_fee          # I = 수수료금액
    print(f"\n✓ 수수료율 오류 복구: row={fix_row}, 수수료율=0.003, 수수료금액={new_fee:,}")

# 거래금액 오류 4건 추가
for i, r in enumerate(amt_err_targets):
    row_no = r['row']
    ws_raw.cell(row_no, 7).value = 0   # G = 거래금액 = 0
    ws_raw.cell(row_no, 9).value = 0   # I = 수수료금액 = 0 (2차 오류 방지)
    v = r['vals']
    print(f"✓ 거래금액 오류 추가 {i+1}: row={row_no} {v[0]} {v[1]} 금액=0")

# 상품유형 오류 2건 추가 (INVALID_PRODS 순서대로)
INVALID_PRODS = ['FX', 'FWD']
for i, r in enumerate(prod_err_targets):
    row_no = r['row']
    ws_raw.cell(row_no, 5).value = INVALID_PRODS[i]   # E = 상품유형
    v = r['vals']
    print(f"✓ 상품유형 오류 추가 {i+1}: row={row_no} {v[0]} {v[1]} → 상품={INVALID_PRODS[i]}")

# ── 검증: 최종 오류 현황 ──────────────────────────────────────────────────────
print("\n=== 최종 오류 검증 ===")
amt_err_rows = []
prod_err_rows = []
fee_rate_err_rows = []

for row_obj in ws_raw.iter_rows(min_row=5, values_only=True):
    if row_obj[0] is None or row_obj[9] != '정상': continue
    date, branch, emp, cust, prod, trade, amt, rate, fee, status = row_obj[:10]
    try: amt_f = float(amt) if amt is not None else None
    except: amt_f = None
    try: rate_f = float(rate) if rate is not None else None
    except: rate_f = None

    if amt_f is None or amt_f <= 0:
        amt_err_rows.append((date, branch, prod))
    if rate_f is not None and (rate_f <= 0 or rate_f > 0.005):
        fee_rate_err_rows.append((date, branch, rate_f))
    if prod not in VALID_PROD:
        prod_err_rows.append((date, branch, prod))

print(f"거래금액 오류: {len(amt_err_rows)}건")
for r in amt_err_rows: print(f"  {r}")
print(f"수수료율 오류: {len(fee_rate_err_rows)}건 (0이어야 함)")
for r in fee_rate_err_rows: print(f"  {r}")
print(f"상품유형 오류: {len(prod_err_rows)}건")
for r in prod_err_rows: print(f"  {r}")

# ──────────────────────────────────────────────────────────────────────────────
# 2. STEP1 재구성: 단일 직원별 집계표
# ──────────────────────────────────────────────────────────────────────────────

# 직원 마스터 로드
master = {}
for row in ws_m.iter_rows(min_row=4, values_only=True):
    if row[0] and str(row[0]).startswith('E'):
        master[row[0]] = (row[1], row[2])

ws1 = wb['결과작성_STEP1_보고용집계']

# 병합 해제 → 일반 셀 초기화
for rng in list(ws1.merged_cells.ranges):
    ws1.unmerge_cells(str(rng))
for row in ws1.iter_rows():
    for cell in row:
        cell.value     = None
        cell.fill      = openpyxl.styles.PatternFill(fill_type=None)
        cell.font      = openpyxl.styles.Font()
        cell.border    = openpyxl.styles.Border()
        cell.alignment = openpyxl.styles.Alignment()

for i in range(1, 50):
    ws1.row_dimensions[i].height = None

# 열 너비 설정
col_widths = {'A':10,'B':12,'C':10,'D':12,'E':18,'F':18,'G':14,'H':4}
for col, w in col_widths.items():
    ws1.column_dimensions[col].width = w

# ── 스타일 ────────────────────────────────────────────────────────────────────
TITLE_FILL  = PatternFill("solid", fgColor="1F3864")
TITLE_FONT  = Font(bold=True, color="FFFFFF", size=14)
SUB_FILL    = PatternFill("solid", fgColor="2F5496")
SUB_FONT    = Font(bold=True, color="FFFFFF", size=10)
HEAD_FILL   = PatternFill("solid", fgColor="D6E4F0")
HEAD_FONT   = Font(bold=True, color="1F3864", size=10)
ID_FILL     = PatternFill("solid", fgColor="F2F6FC")
INPUT_FILL  = PatternFill("solid", fgColor="FFFDE7")  # 연노랑 = 학생 입력칸
THIN        = Side(style="thin", color="BFBFBF")
MED         = Side(style="medium", color="4472C4")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED, right=MED, top=MED, bottom=MED)

def set_cell(ws, row, col, value=None, fill=None, font=None, align=None, border=None, fmt=None):
    c = ws.cell(row, col)
    if value is not None: c.value = value
    if fill:   c.fill  = fill
    if font:   c.font  = font
    if align:  c.alignment = align
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left',   vertical='center')
right  = Alignment(horizontal='right',  vertical='center')

# ── 행 1: 타이틀 ─────────────────────────────────────────────────────────────
ws1.merge_cells('A1:H1')
set_cell(ws1, 1, 1,
    value='STEP 1 · 직원별 거래 현황 집계표',
    fill=TITLE_FILL, font=TITLE_FONT, align=center)
ws1.row_dimensions[1].height = 36

# ── 행 2: 설명 ───────────────────────────────────────────────────────────────
ws1.merge_cells('A2:H2')
set_cell(ws1, 2, 1,
    value='📋  직원마스터에서 직원명·지점코드를 조회하고, 거래원장_RAW에서 집계하세요 (처리상태=정상 기준)',
    fill=PatternFill("solid", fgColor="EBF1F5"),
    font=Font(italic=True, color="375623", size=9),
    align=Alignment(horizontal='left', vertical='center'))
ws1.row_dimensions[2].height = 22

# ── 행 3: 공백 ───────────────────────────────────────────────────────────────
ws1.row_dimensions[3].height = 6

# ── 행 4: 컬럼 헤더 ──────────────────────────────────────────────────────────
headers = [
    (1, '직원 ID'),
    (2, '직원명\n(VLOOKUP)'),
    (3, '지점코드\n(VLOOKUP)'),
    (4, '정상\n거래건수'),
    (5, '정상\n거래금액'),
    (6, '정상\n수수료합계'),
    (7, '취소·오류\n건수'),
]
for col, h in headers:
    set_cell(ws1, 4, col, value=h,
             fill=HEAD_FILL, font=HEAD_FONT, align=center, border=THIN_BORDER)
ws1.row_dimensions[4].height = 34

# ── 행 5~24: 직원별 데이터 행 (E001~E020) ─────────────────────────────────────
EMPLOYEES = [f'E{str(i).zfill(3)}' for i in range(1,21)]

for i, emp_id in enumerate(EMPLOYEES):
    r = 5 + i
    ws1.row_dimensions[r].height = 20

    # A: 직원ID (사전 입력)
    set_cell(ws1, r, 1, value=emp_id,
             fill=ID_FILL, font=Font(bold=True, color="2F5496", size=9),
             align=center, border=THIN_BORDER)

    # B~G: 학생 입력칸
    for col in range(2, 8):
        set_cell(ws1, r, col,
                 fill=INPUT_FILL,
                 font=Font(color="333333", size=9),
                 align=right if col >= 4 else center,
                 border=THIN_BORDER)

# ── 행 25: 합계 행 ────────────────────────────────────────────────────────────
ws1.row_dimensions[25].height = 22
set_cell(ws1, 25, 1, value='합 계',
         fill=PatternFill("solid", fgColor="D6E4F0"),
         font=Font(bold=True, color="1F3864", size=9),
         align=center, border=THIN_BORDER)
for col in range(2, 8):
    c = ws1.cell(25, col)
    c.fill   = PatternFill("solid", fgColor="D6E4F0")
    c.font   = Font(bold=True, color="1F3864", size=9)
    c.border = THIN_BORDER
    c.alignment = right if col >= 4 else center

# ── 행 27~: 힌트 박스 ─────────────────────────────────────────────────────────
ws1.merge_cells('A27:H27')
set_cell(ws1, 27, 1,
    value='💡  AI 프롬프트 예시',
    fill=PatternFill("solid", fgColor="FFF3CD"),
    font=Font(bold=True, color="856404", size=10),
    align=Alignment(horizontal='left', vertical='center'))
ws1.row_dimensions[27].height = 22

hints = [
    (28, '"A5셀(E001)의 직원명을 직원마스터 시트에서 찾아서 B5에 넣는 VLOOKUP 수식 알려줘"'),
    (29, '"거래원장_RAW에서 A5셀 직원의 정상 거래건수를 세는 COUNTIFS 수식 알려줘"'),
    (30, '"위와 같은 방식으로 D5:G5 범위를 한번에 채울 수 있도록 수식 만들어줘"'),
]
for row_no, hint_text in hints:
    ws1.merge_cells(f'A{row_no}:H{row_no}')
    set_cell(ws1, row_no, 1,
        value=hint_text,
        fill=PatternFill("solid", fgColor="FFFDE7"),
        font=Font(italic=True, color="856404", size=8),
        align=Alignment(horizontal='left', vertical='center', wrap_text=True))
    ws1.row_dimensions[row_no].height = 18

print("\n✓ STEP1 재구성 완료")

# ──────────────────────────────────────────────────────────────────────────────
# 3. STEP2 수정: 2개 오류유형만
# ──────────────────────────────────────────────────────────────────────────────
ws2 = wb['결과작성_STEP2_오류찾기']

for rng in list(ws2.merged_cells.ranges):
    ws2.unmerge_cells(str(rng))
for row in ws2.iter_rows():
    for cell in row:
        cell.value     = None
        cell.fill      = openpyxl.styles.PatternFill(fill_type=None)
        cell.font      = openpyxl.styles.Font()
        cell.border    = openpyxl.styles.Border()
        cell.alignment = openpyxl.styles.Alignment()

for i in range(1, 30):
    ws2.row_dimensions[i].height = None

# 열 너비
for col, w in [('A',18),('B',38),('C',10),('D',32)]:
    ws2.column_dimensions[col].width = w

# ── 타이틀 ───────────────────────────────────────────────────────────────────
ws2.merge_cells('A1:D1')
set_cell(ws2, 1, 1,
    value='STEP 2 · 이상 거래 탐지',
    fill=TITLE_FILL, font=TITLE_FONT, align=center)
ws2.row_dimensions[1].height = 36

ws2.merge_cells('A2:D2')
set_cell(ws2, 2, 1,
    value='📋  거래원장_RAW에서 처리상태=정상이지만 실제 값이 잘못된 행을 찾아 K열에 오류유형을 표시하고, 아래 표에 발견 건수를 입력하세요.',
    fill=PatternFill("solid", fgColor="EBF1F5"),
    font=Font(italic=True, color="375623", size=9),
    align=Alignment(horizontal='left', vertical='center', wrap_text=True))
ws2.row_dimensions[2].height = 30

ws2.row_dimensions[3].height = 8

# ── 헤더 ─────────────────────────────────────────────────────────────────────
s2_headers = ['오류 유형', '판단 기준', '발견 건수', '비고']
for col, h in enumerate(s2_headers, 1):
    set_cell(ws2, 4, col, value=h,
             fill=HEAD_FILL, font=HEAD_FONT, align=center, border=THIN_BORDER)
ws2.row_dimensions[4].height = 26

# ── 오류 유형 2개 ─────────────────────────────────────────────────────────────
ERR_TYPES = [
    ('거래금액 오류',  '처리상태=정상  AND  거래금액 ≤ 0',     '허용값: 0 초과'),
    ('상품유형 오류',  '처리상태=정상  AND  상품유형이 주식/펀드/채권/ELS/RP 외', '허용값: 주식, 펀드, 채권, ELS, RP'),
]
for i, (etype, criteria, note) in enumerate(ERR_TYPES):
    r = 5 + i
    ws2.row_dimensions[r].height = 26
    set_cell(ws2, r, 1, value=etype,
             fill=PatternFill("solid", fgColor="F2DEDE"),
             font=Font(bold=True, color="8B0000", size=9),
             align=center, border=THIN_BORDER)
    set_cell(ws2, r, 2, value=criteria,
             fill=PatternFill("solid", fgColor="FFF9F9"),
             font=Font(color="333333", size=9),
             align=left, border=THIN_BORDER)
    set_cell(ws2, r, 3, value=None,   # 학생 입력
             fill=INPUT_FILL,
             font=Font(bold=True, size=10),
             align=center, border=THIN_BORDER)
    set_cell(ws2, r, 4, value=note,
             fill=PatternFill("solid", fgColor="FFF9F9"),
             font=Font(color="595959", size=9, italic=True),
             align=left, border=THIN_BORDER)

# ── 공백 + K열 안내 ──────────────────────────────────────────────────────────
ws2.row_dimensions[8].height = 10
ws2.merge_cells('A9:D9')
set_cell(ws2, 9, 1,
    value='📌  오류를 찾으면 거래원장_RAW의 K열(오류유형_작성)에 오류유형명을 입력하세요.',
    fill=PatternFill("solid", fgColor="FFF3CD"),
    font=Font(bold=True, color="856404", size=9),
    align=Alignment(horizontal='left', vertical='center'))
ws2.row_dimensions[9].height = 22

# ── AI 프롬프트 힌트 ──────────────────────────────────────────────────────────
ws2.merge_cells('A11:D11')
set_cell(ws2, 11, 1,
    value='💡  AI 프롬프트 예시',
    fill=PatternFill("solid", fgColor="FFF3CD"),
    font=Font(bold=True, color="856404", size=10),
    align=Alignment(horizontal='left', vertical='center'))
ws2.row_dimensions[11].height = 22

s2_hints = [
    (12, '"거래원장_RAW에서 J열(처리상태)이 정상인데 G열(거래금액)이 0 이하인 행을 찾아 K열에 \'거래금액 오류\'를 표시하는 수식 알려줘"'),
    (13, '"처리상태=정상인데 E열(상품유형)이 주식/펀드/채권/ELS/RP가 아닌 행을 찾아 K열에 \'상품유형 오류\'를 표시하는 IF+ISNUMBER+MATCH 수식 알려줘"'),
    (14, '"위 두 조건을 합쳐서 K열에 한꺼번에 표시하는 수식 만들어줘"'),
]
for row_no, hint_text in s2_hints:
    ws2.merge_cells(f'A{row_no}:D{row_no}')
    set_cell(ws2, row_no, 1,
        value=hint_text,
        fill=PatternFill("solid", fgColor="FFFDE7"),
        font=Font(italic=True, color="856404", size=8),
        align=Alignment(horizontal='left', vertical='center', wrap_text=True))
    ws2.row_dimensions[row_no].height = 20

print("✓ STEP2 재구성 완료")

# ──────────────────────────────────────────────────────────────────────────────
# 저장
# ──────────────────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f"\n✅ 저장 완료: {DEST}")
