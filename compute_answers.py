"""v3_fixed 파일에서 STEP1/STEP2 정답 계산"""
import sys
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

PATH = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_fixed.xlsx'
wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb['거래원장_RAW']

# 5행부터 데이터
rows = list(ws.iter_rows(min_row=5, values_only=True))
# 컬럼: A날짜 B지점 C직원 D고객 E상품 F거래유형 G거래금액 H수수료율 I수수료금액 J처리상태

VALID_PROD  = {'주식','펀드','채권','ELS','RP'}
VALID_TRADE = {'매수','매도','가입','해지'}

# 직원마스터 로드
ws_m = wb['직원마스터']
master = {}
for row in ws_m.iter_rows(min_row=4, values_only=True):
    if row[0] and row[1]:
        master[row[0]] = (row[1], row[2])  # 직원ID -> (직원명, 지점코드)

print("=== STEP1 정답 ===\n")

# ── 지점별 집계 ──
branch_normal_cnt   = defaultdict(int)
branch_normal_amt   = defaultdict(int)
branch_normal_fee   = defaultdict(int)
branch_err_cnt      = defaultdict(int)

for r in rows:
    branch = r[1]
    status = r[9]
    amt = int(r[6]) if r[6] else 0
    fee = int(r[8]) if r[8] else 0
    if not branch: continue
    if status == '정상':
        branch_normal_cnt[branch] += 1
        branch_normal_amt[branch] += amt
        branch_normal_fee[branch] += fee
    else:
        branch_err_cnt[branch] += 1

print("지점별:")
for b in ['B001','B002','B003','B004','B005']:
    print(f"  {b}: 정상건수={branch_normal_cnt[b]}, 거래금액={branch_normal_amt[b]:,}, 수수료={branch_normal_fee[b]:,}, 취소오류={branch_err_cnt[b]}")

# ── 상품유형별 집계 ──
prod_normal_cnt = defaultdict(int)
prod_normal_amt = defaultdict(int)
prod_normal_fee = defaultdict(int)
total_normal_amt = sum(int(r[6]) for r in rows if r[9]=='정상' and r[6])

for r in rows:
    prod = r[4]
    status = r[9]
    amt = int(r[6]) if r[6] else 0
    fee = int(r[8]) if r[8] else 0
    if status == '정상' and prod:
        prod_normal_cnt[prod] += 1
        prod_normal_amt[prod] += amt
        prod_normal_fee[prod] += fee

print(f"\n상품유형별 (전체 정상 거래금액 합계={total_normal_amt:,}):")
for p in ['주식','펀드','채권','ELS','RP']:
    ratio = prod_normal_amt[p]/total_normal_amt*100 if total_normal_amt else 0
    print(f"  {p}: 건수={prod_normal_cnt[p]}, 금액={prod_normal_amt[p]:,}, 수수료={prod_normal_fee[p]:,}, 비중={ratio:.1f}%")

# ── 직원별 처리 현황 ──
emp_normal_cnt  = defaultdict(int)
emp_normal_fee  = defaultdict(int)
emp_err_cnt     = defaultdict(int)

for r in rows:
    emp = r[2]
    status = r[9]
    fee = int(r[8]) if r[8] else 0
    if not emp: continue
    if status == '정상':
        emp_normal_cnt[emp] += 1
        emp_normal_fee[emp] += fee
    else:
        emp_err_cnt[emp] += 1

print("\n직원별:")
for e in [f'E{str(i).zfill(3)}' for i in range(1,21)]:
    name, branch = master.get(e, ('?','?'))
    print(f"  {e} {name} {branch}: 정상건수={emp_normal_cnt[e]}, 수수료={emp_normal_fee[e]:,}, 취소오류={emp_err_cnt[e]}")

print("\n\n=== STEP2 정답 ===\n")
normal = [r for r in rows if r[9] == '정상']

trade_vals = set(r[5] for r in rows if r[5])

err1=[]; err2=[]; err3=[]; err4=[]; err5=[]
for r in normal:
    date,branch,emp,cust,prod,trade,amt,rate,fee,status = r[:10]
    try: amt_f = float(amt) if amt is not None else None
    except: amt_f = None
    try: rate_f = float(rate) if rate is not None else None
    except: rate_f = None
    try: fee_i = int(fee) if fee is not None else None
    except: fee_i = None

    if amt_f is None or amt_f <= 0: err1.append(r)
    if rate_f is None or rate_f <= 0 or rate_f > 0.005: err2.append(r)
    if amt_f and rate_f and fee_i is not None:
        if abs(round(amt_f * rate_f) - fee_i) > 1: err3.append(r)
    if prod not in VALID_PROD: err4.append(r)
    if trade not in VALID_TRADE: err5.append(r)

print(f"거래금액 오류: {len(err1)}건")
for r in err1: print(f"  행: {r[0]} {r[1]} {r[4]} 거래금액={r[6]}")

print(f"수수료율 오류: {len(err2)}건")
for r in err2: print(f"  행: {r[0]} {r[1]} {r[4]} 수수료율={r[7]}")

print(f"수수료금액 오류: {len(err3)}건")
for r in err3: print(f"  행: {r[0]} {r[1]} {r[4]}")

print(f"상품유형 오류: {len(err4)}건")
for r in err4: print(f"  행: {r[0]} {r[1]} 상품={r[4]}")

print(f"거래유형 오류: {len(err5)}건")
for r in err5: print(f"  행: {r[0]} {r[1]} 거래유형={r[5]}")

# K열 정답 (오류 3행의 행번호 & 기대 오류유형)
print("\n=== K열 정답 행 ===")
for r in rows:
    if r[9] == '정상':
        date,branch,emp,cust,prod,trade,amt,rate,fee,status = r[:10]
        try: amt_f = float(amt) if amt else None
        except: amt_f = None
        try: rate_f = float(rate) if rate else None
        except: rate_f = None
        try: fee_i = int(fee) if fee else None
        except: fee_i = None
        
        errs = []
        if amt_f is None or amt_f <= 0: errs.append('거래금액 오류')
        elif rate_f is None or rate_f <= 0 or rate_f > 0.005: errs.append('수수료율 오류')
        elif amt_f and rate_f and fee_i is not None and abs(round(amt_f*rate_f)-fee_i)>1: errs.append('수수료금액 오류')
        elif prod not in VALID_PROD: errs.append('상품유형 오류')
        elif trade not in VALID_TRADE: errs.append('거래유형 오류')
        
        if errs:
            print(f"  날짜={date} 지점={branch} 직원={emp} 고객={cust} → K열='{errs[0]}'")
