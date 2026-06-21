"""
v3_fixed 학생 파일 → 강사용 정답 파일 생성
  - 결과작성_STEP1_보고용집계  : 모든 빈칸 채움
  - 결과작성_STEP2_오류찾기    : 발견건수 C5:C9 채움
  - 거래원장_RAW               : K열 3개 오류행 채움
  - 시트탭 색상 변경 (강사용 표시)
  - 최상단 공지 행 삽입
"""
import shutil, sys, re
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC  = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_fixed.xlsx'
DEST = r'C:\Users\jay\Downloads\hana_ai_excel_practice_instructor_v3.xlsx'

shutil.copy2(SRC, DEST)
wb = openpyxl.load_workbook(DEST)

# ── 원본 데이터에서 정답 계산 ─────────────────────────────────────────────────
ws_raw = wb['거래원장_RAW']
ws_m   = wb['직원마스터']

VALID_PROD  = {'주식','펀드','채권','ELS','RP'}
VALID_TRADE = {'매수','매도','가입','해지'}

# 직원 마스터
master = {}   # empID -> (name, branch)
for row in ws_m.iter_rows(min_row=4, values_only=True):
    if row[0] and str(row[0]).startswith('E'):
        master[row[0]] = (row[1], row[2])

# RAW 데이터 순회
rows = list(ws_raw.iter_rows(min_row=5, values_only=True))

branch_normal_cnt = defaultdict(int)
branch_normal_amt = defaultdict(int)
branch_normal_fee = defaultdict(int)
branch_err_cnt    = defaultdict(int)

prod_normal_cnt   = defaultdict(int)
prod_normal_amt   = defaultdict(int)
prod_normal_fee   = defaultdict(int)

emp_normal_cnt    = defaultdict(int)
emp_normal_fee    = defaultdict(int)
emp_err_cnt       = defaultdict(int)

for r in rows:
    date, branch, emp, cust, prod, trade, amt, rate, fee, status = r[:10]
    if not branch: continue
    amt_i = int(amt) if amt else 0
    fee_i = int(fee) if fee else 0
    if status == '정상':
        branch_normal_cnt[branch] += 1
        branch_normal_amt[branch] += amt_i
        branch_normal_fee[branch] += fee_i
        if prod:
            prod_normal_cnt[prod] += 1
            prod_normal_amt[prod] += amt_i
            prod_normal_fee[prod] += fee_i
        if emp:
            emp_normal_cnt[emp] += 1
            emp_normal_fee[emp] += fee_i
    else:
        branch_err_cnt[branch] += 1
        if emp:
            emp_err_cnt[emp] += 1

total_normal_amt = sum(branch_normal_amt.values())
BRANCHES = ['B001','B002','B003','B004','B005']
PRODUCTS = ['주식','펀드','채권','ELS','RP']
EMPLOYEES= [f'E{str(i).zfill(3)}' for i in range(1,21)]

print("── 정답 요약 ──")
for b in BRANCHES:
    print(f"{b}: 건수={branch_normal_cnt[b]}, 금액={branch_normal_amt[b]:,}, 수수료={branch_normal_fee[b]:,}, 취소오류={branch_err_cnt[b]}")
for p in PRODUCTS:
    ratio = prod_normal_amt[p]/total_normal_amt if total_normal_amt else 0
    print(f"{p}: 건수={prod_normal_cnt[p]}, 금액={prod_normal_amt[p]:,}, 수수료={prod_normal_fee[p]:,}, 비중={ratio:.4f}")
for e in EMPLOYEES:
    name, branch = master.get(e, ('?','?'))
    print(f"{e} {name} {branch}: 정상건수={emp_normal_cnt[e]}, 수수료={emp_normal_fee[e]:,}, 취소오류={emp_err_cnt[e]}")

# ── 스타일 도우미 ──────────────────────────────────────────────────────────────
ANSWER_FILL = PatternFill("solid", fgColor="FFF2CC")   # 연한 노란색
ANS_FONT    = Font(bold=True, color="7F4600")
HEADER_FILL = PatternFill("solid", fgColor="FF6B35")   # 강사용 헤더 주황
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)

def write_ans(ws, cell_ref, value, fmt=None):
    c = ws[cell_ref]
    c.value = value
    c.fill  = ANSWER_FILL
    c.font  = ANS_FONT
    if fmt:
        c.number_format = fmt

# ── STEP1 작성 ────────────────────────────────────────────────────────────────
ws1 = wb['결과작성_STEP1_보고용집계']

for i, b in enumerate(BRANCHES):
    r = 5 + i
    write_ans(ws1, f'B{r}', branch_normal_cnt[b], '0')
    write_ans(ws1, f'C{r}', branch_normal_amt[b], '#,##0')
    write_ans(ws1, f'D{r}', branch_normal_fee[b], '#,##0')
    write_ans(ws1, f'E{r}', branch_err_cnt[b],    '0')

for i, p in enumerate(PRODUCTS):
    r = 5 + i
    ratio = prod_normal_amt[p] / total_normal_amt if total_normal_amt else 0
    write_ans(ws1, f'I{r}', prod_normal_cnt[p],  '0')
    write_ans(ws1, f'J{r}', prod_normal_amt[p],  '#,##0')
    write_ans(ws1, f'K{r}', prod_normal_fee[p],  '#,##0')
    write_ans(ws1, f'L{r}', ratio,               '0.0%')

for i, e in enumerate(EMPLOYEES):
    r = 13 + i       # row 12 = 직원ID 헤더, row 13부터 E001 데이터
    name, branch = master.get(e, ('?','?'))
    write_ans(ws1, f'B{r}', name,                 '@')
    write_ans(ws1, f'C{r}', branch,               '@')
    write_ans(ws1, f'D{r}', emp_normal_cnt[e],    '0')
    write_ans(ws1, f'E{r}', emp_normal_fee[e],    '#,##0')
    write_ans(ws1, f'F{r}', emp_err_cnt[e],       '0')

print("✓ STEP1 완료")

# ── STEP2 요약 건수 작성 ──────────────────────────────────────────────────────
ws2 = wb['결과작성_STEP2_오류찾기']

# STEP2에서 직접 계산: 정상 행 기준 오류 체크
err_counts = {t: 0 for t in ['거래금액 오류','수수료율 오류','수수료금액 오류','상품유형 오류','거래유형 오류']}

for r in rows:
    date, branch, emp, cust, prod, trade, amt, rate, fee, status = r[:10]
    if status != '정상': continue
    try: amt_f = float(amt) if amt is not None else None
    except: amt_f = None
    try: rate_f = float(rate) if rate is not None else None
    except: rate_f = None
    try: fee_i2 = int(fee) if fee is not None else None
    except: fee_i2 = None

    if amt_f is None or amt_f <= 0:
        err_counts['거래금액 오류'] += 1
    elif rate_f is None or rate_f <= 0 or rate_f > 0.005:
        err_counts['수수료율 오류'] += 1
    elif amt_f and rate_f and fee_i2 is not None and abs(round(amt_f * rate_f) - fee_i2) > 1:
        err_counts['수수료금액 오류'] += 1
    elif prod not in VALID_PROD:
        err_counts['상품유형 오류'] += 1
    elif trade not in VALID_TRADE:
        err_counts['거래유형 오류'] += 1

err_order = ['거래금액 오류','수수료율 오류','수수료금액 오류','상품유형 오류','거래유형 오류']
for i, t in enumerate(err_order):
    write_ans(ws2, f'C{5+i}', err_counts[t], '0')
    print(f"  {t}: {err_counts[t]}건 → C{5+i}")

print("✓ STEP2 완료")

# ── 거래원장_RAW K열 작성 ──────────────────────────────────────────────────────
# 오류 행 식별
ERR_ROWS = {
    (20250104, 'B004', 'E004', 'C066'): '거래금액 오류',
    (20250109, 'B002', 'E002', 'C021'): '수수료율 오류',
    (20250122, 'B002', 'E017', 'C065'): '상품유형 오류',
}

K_FILL = PatternFill("solid", fgColor="FFE4B5")   # 연한 오렌지
K_FONT = Font(bold=True, color="8B0000")

for row_idx, r in enumerate(rows):
    date, branch, emp, cust = r[0], r[1], r[2], r[3]
    key = (date, branch, emp, cust)
    if key in ERR_ROWS:
        excel_row = row_idx + 5          # 데이터는 5행부터
        c = ws_raw.cell(excel_row, 11)   # K열 = 11번째
        c.value = ERR_ROWS[key]
        c.fill  = K_FILL
        c.font  = K_FONT
        print(f"  RAW K{excel_row} = '{ERR_ROWS[key]}'  ({date} {branch})")

print("✓ 거래원장_RAW K열 완료")

# ── 시트탭 색상 (강사용 강조) ─────────────────────────────────────────────────
for sn in ['결과작성_STEP1_보고용집계','결과작성_STEP2_오류찾기']:
    wb[sn].sheet_properties.tabColor = "FF6B35"

ws_raw.sheet_properties.tabColor = "C00000"

# ── 강사용 주석 행 삽입 (STEP1 맨 위) ────────────────────────────────────────
# 기존 행 1 위에 새 행 삽입
ws1.insert_rows(1)
ws1.merge_cells('A1:M1')
hdr = ws1['A1']
hdr.value = '★ 강사용 정답 파일 — 배포 금지 ★'
hdr.fill  = HEADER_FILL
hdr.font  = HEADER_FONT
hdr.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 24

ws2.insert_rows(1)
ws2.merge_cells('A1:F1')
hdr2 = ws2['A1']
hdr2.value = '★ 강사용 정답 파일 — 배포 금지 ★'
hdr2.fill  = HEADER_FILL
hdr2.font  = HEADER_FONT
hdr2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 24

# ── 저장 ──────────────────────────────────────────────────────────────────────
wb.save(DEST)
print(f"\n✅ 저장 완료: {DEST}")
