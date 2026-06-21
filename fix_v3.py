"""
V3 파일 수정:
  - 오류 3건만 남기기 (거래금액=0, 수수료율=1.5%, ETF)
  - 나머지 불필요한 오류 수치 정상화
  - 결과작성_STEP2 거래유형 오류 판단기준 추가
  - 프롬프트는 원본 유지 (해지가 올바른 데이터 값)
"""
import sys, shutil
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

SRC = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_simplified_errorcheck.xlsx'
DST = r'C:\Users\jay\Downloads\hana_ai_excel_practice_student_v3_fixed.xlsx'
shutil.copy2(SRC, DST)

wb = openpyxl.load_workbook(DST)

# ─────────────────────────────────────────────────────────────
# 1. 거래원장_RAW: 불필요한 오류 수치 정상화
#    유지: 거래금액=0 / 수수료율=1.5% / ETF 상품유형 (각 1건씩 3개)
# ─────────────────────────────────────────────────────────────
ws = wb['거래원장_RAW']

# 컬럼 인덱스 (1-based 열번호, openpyxl row[n-1])
# A=날짜(1) B=지점(2) C=직원(3) D=고객(4) E=상품(5) F=거래유형(6)
# G=거래금액(7) H=수수료율(8) I=수수료금액(9) J=처리상태(10)

FIXES = {
    # 거래금액=0 행: 수수료금액도 0으로 맞춤 (0×rate=0이 맞음)
    (20250104, 'B004', 'E004', 'C066'): {9: 0},

    # 수수료율=1.5% 행: 수수료금액을 1.5% 기준값으로 정정 (오류는 수수료율만)
    # 472500000 × 0.015 = 7,087,500
    (20250109, 'B002', 'E002', 'C021'): {9: 7087500},

    # 수수료금액 불일치 행 (493900000×0.0012=592680, 실제=670457) → 정정
    (20250117, 'B005', 'E015', 'C007'): {9: 592680},

    # 거래유형=정정 행 → 매도로 정정 (정정은 허용값 아님, 하지만 3개 오류 밖이므로 제거)
    (20250129, 'B001', 'E001', 'C040'): {6: '매도'},

    # 수수료금액=-5000 행 → 103900000×0.001=103,900
    (20250224, 'B002', 'E012', 'C035'): {9: 103900},

    # 거래금액=-1,000,000 행 → 정상값 복원 (85,000,000), 수수료금액 정정
    # 85000000 × 0.0029 = 246,500
    (20250306, 'B001', 'E001', 'C079'): {7: 85000000, 9: 246500},

    # 수수료율=None 행 → 0.0015, 수수료금액=16700000×0.0015=25,050
    (20250319, 'B003', 'E008', 'C071'): {8: 0.0015, 9: 25050},
}

modified = []
for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
    key = (row[0].value, row[1].value, row[2].value, row[3].value)
    if key in FIXES:
        for col_idx, new_val in FIXES[key].items():
            row[col_idx - 1].value = new_val
        modified.append(f"  수정: {key} → {FIXES[key]}")

print(f"RAW 수정 완료: {len(modified)}행")
for m in modified:
    print(m)


# ─────────────────────────────────────────────────────────────
# 2. 결과작성_STEP2_오류찾기: 거래유형 오류 판단기준 추가
# ─────────────────────────────────────────────────────────────
ws_step2 = wb['결과작성_STEP2_오류찾기']

for row in ws_step2.iter_rows(min_col=1, max_col=1):
    cell_a = row[0]
    if cell_a.value and '거래유형' in str(cell_a.value):
        r = cell_a.row
        to_unmerge = []
        for merge in ws_step2.merged_cells.ranges:
            if merge.min_row <= r <= merge.max_row:
                if (merge.min_col <= 2 <= merge.max_col) or (merge.min_col <= 4 <= merge.max_col):
                    to_unmerge.append(str(merge))
        for rng in to_unmerge:
            try:
                ws_step2.unmerge_cells(rng)
            except ValueError:
                pass
        ws_step2.cell(r, 2).value = '처리상태=정상  AND  거래유형이 매수/매도/가입/해지 외'
        ws_step2.cell(r, 4).value = '허용값: 매수/매도/가입/해지'
        print(f"\nSTEP2 거래유형 오류 판단기준 추가 (행 {r})")
        break

wb.save(DST)


# ─────────────────────────────────────────────────────────────
# 3. 검증: 수정 후 정상 행 오류 재카운트 (정답은 3건이어야 함)
# ─────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(DST, data_only=True)
ws_r = wb2['거래원장_RAW']
rows = list(ws_r.iter_rows(min_row=5, values_only=True))
normal = [r for r in rows if r[9] == '정상']

VALID_PROD  = {'주식','펀드','채권','ELS','RP'}
# 거래유형 실제 값 확인 (첫 해지/가입 행 출력)
trade_vals = set(r[5] for r in rows if r[5])
print(f"\n거래유형 전체 값: {trade_vals}")
VALID_TRADE = set(t for t in trade_vals if t not in {'정정'})  # 비정상 1개 제외한 나머지
print(f"정상 거래유형으로 간주: {VALID_TRADE}")

errors = []
for r in normal:
    date,branch,emp,cust,prod,trade,amt,rate,fee,status = r[:10]
    try: amt_f = float(amt) if amt is not None else None
    except: amt_f = None
    try: rate_f = float(rate) if rate is not None else None
    except: rate_f = None
    try: fee_i = int(fee) if fee is not None else None
    except: fee_i = None

    errs = []
    if amt_f is None or amt_f <= 0:
        errs.append('거래금액오류')
    if rate_f is None or rate_f <= 0 or rate_f > 0.005:
        errs.append('수수료율오류')
    if amt_f and rate_f and fee_i is not None:
        if abs(round(amt_f * rate_f) - fee_i) > 1:
            errs.append('수수료금액불일치')
    if prod not in VALID_PROD:
        errs.append('상품유형오류')
    if trade not in VALID_TRADE:
        errs.append('거래유형오류')
    if errs:
        errors.append((date, branch, prod, trade, errs))

print(f"\n=== 최종 오류 현황 ===")
print(f"정상 행: {len(normal)}건 / 오류 감지: {len(errors)}건 (목표: 3건)")
for e in errors:
    print(f"  {e[0]} {e[1]} {e[2]} {e[3]}: {e[4]}")

if len(errors) == 3:
    print("\n✓ 정답! 오류 3건 정확히 심어졌습니다.")
else:
    print(f"\n✗ 오류 건수 불일치. 추가 확인 필요.")

print(f"\n저장 완료 → {DST}")
